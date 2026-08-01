# -*- coding: utf-8 -*-

from dateutil.relativedelta import relativedelta

from odoo import api, fields, models


MIEN_COLORS = {
    "BAC": "#ef4444",
    "TRUNG": "#22c55e",
    "NAM": "#3b82f6",
    "DTT": "#a855f7",
    "VP": "#64748b",
}
MIEN_FALLBACK_COLORS = ["#ef4444", "#22c55e", "#3b82f6", "#a855f7", "#f59e0b", "#0d9488"]


class PhanHeDashboard(models.AbstractModel):
    _name = "phan.he.dashboard"
    _description = "Dashboard quản lý dịch vụ"

    @api.model
    def get_dashboard_data(self, filters=None):
        filters = filters or {}
        today = fields.Date.context_today(self)
        now = fields.Datetime.context_timestamp(self, fields.Datetime.now())

        year = today.year
        if filters.get("date_from"):
            year = fields.Date.to_date(filters["date_from"]).year
        date_from = fields.Date.to_date(filters.get("date_from") or f"{year}-01-01")
        date_to = fields.Date.to_date(filters.get("date_to") or f"{year}-12-31")
        year = date_from.year

        mien_id = int(filters["mien_id"]) if filters.get("mien_id") else False
        area_id = int(filters["area_id"]) if filters.get("area_id") else False
        emp_id = int(filters["employee_id"]) if filters.get("employee_id") else False

        Service = self.env["phan.he.service"]
        Payment = self.env["phan.he.payment"]
        Mien = self.env["phan.he.mien"]

        s_domain = [("active", "=", True)]
        service_type_code = (filters.get("service_type_code") or "").strip().lower()
        if service_type_code:
            stype = self.env["phan.he.service.type"].search(
                [("code", "=", service_type_code)], limit=1
            )
            if stype:
                s_domain.append(("service_type_id", "=", stype.id))
        if mien_id:
            s_domain.append(("mien_id", "=", mien_id))
        if area_id:
            s_domain.append(("area_id", "=", area_id))
        if emp_id:
            s_domain.append(("store_id.responsible_id", "=", emp_id))

        services = Service.search(s_domain)
        miens = Mien.search([("active", "=", True)], order="sequence, name")
        mien_meta = []
        for i, m in enumerate(miens):
            color = MIEN_COLORS.get((m.code or "").upper()) or MIEN_FALLBACK_COLORS[i % len(MIEN_FALLBACK_COLORS)]
            short = (m.name or "").replace("Miền ", "").strip() or m.code or m.name
            mien_meta.append({
                "id": m.id,
                "code": m.code or "",
                "name": m.name or "",
                "short": short,
                "color": color,
            })

        # --- Chi phí tháng theo miền ---
        # Chỉ cộng cước tháng HĐ còn hiệu lực trong tháng VÀ thời gian còn lại ≤ 30 ngày
        def services_in_month(recs, m_start, m_end):
            return recs.filtered(
                lambda s: s.state not in ("cancel",)
                and (not s.date_start or s.date_start <= m_end)
                and (not s.date_end or s.date_end >= m_start)
            )

        def cost_by_mien(recs, m_start, m_end):
            """Tổng cước tháng: HĐ hiệu lực trong tháng và còn ≤ 30 ngày (tính tới cuối tháng / hôm nay)."""
            active = services_in_month(recs, m_start, m_end)
            as_of = m_end if m_end <= today else today
            result = {m["id"]: 0.0 for m in mien_meta}
            for svc in active:
                if not svc.date_end:
                    continue
                if (svc.date_end - as_of).days > 30:
                    continue
                mid = svc.mien_id.id
                if mid in result:
                    result[mid] += svc.contract_amount or 0.0
            return result

        cur_month_start = today.replace(day=1)
        cur_month_end = cur_month_start + relativedelta(months=1, days=-1)
        prev_month_start = cur_month_start - relativedelta(months=1)
        prev_month_end = cur_month_start - relativedelta(days=1)

        cur_map = cost_by_mien(services, cur_month_start, cur_month_end)
        prev_map = cost_by_mien(services, prev_month_start, prev_month_end)

        region_kpis = []
        for m in mien_meta:
            cur = cur_map.get(m["id"], 0.0)
            prev = prev_map.get(m["id"], 0.0)
            region_kpis.append({
                **m,
                "amount": cur,
                "delta": self._delta_pct(cur, prev),
                "title": f"CHI PHÍ {m['name'].upper()}",
            })

        # --- Cảnh báo ---
        soon30 = today + relativedelta(days=30)
        expire_soon_svcs = services.filtered(
            lambda s: s.state == "active" and s.date_end and today <= s.date_end <= soon30
        )
        expired_svcs = services.filtered(
            lambda s: s.date_end and s.date_end < today and s.state not in ("cancel", "liquidated")
        )
        due_soon_pays = Payment.search([
            ("payment_state", "in", ("due_soon", "pending", "not_due")),
            ("service_id", "in", services.ids or [0]),
            ("date_due", "!=", False),
            ("date_due", ">=", today),
            ("date_due", "<=", soon30),
        ])

        def breakdown_services(recs):
            counts = {m["id"]: 0 for m in mien_meta}
            for rec in recs:
                mid = rec.mien_id.id
                if mid in counts:
                    counts[mid] += 1
            return [
                {"id": m["id"], "short": m["short"], "count": counts[m["id"]], "color": m["color"]}
                for m in mien_meta
            ]

        def breakdown_payments(pays):
            counts = {m["id"]: 0 for m in mien_meta}
            for pay in pays:
                mid = pay.service_id.mien_id.id
                if mid in counts:
                    counts[mid] += 1
            return [
                {"id": m["id"], "short": m["short"], "count": counts[m["id"]], "color": m["color"]}
                for m in mien_meta
            ]

        alert_cards = [
            {
                "id": "expire_soon",
                "level": "warn",
                "icon": "fa-exclamation-triangle",
                "title": "SẮP HẾT HẠN (30 Ngày)",
                "count": len(expire_soon_svcs),
                "unit": "Hợp đồng",
                "breakdown": breakdown_services(expire_soon_svcs),
                "action": "lug_phan_he.action_phan_he_service_expire_soon",
            },
            {
                "id": "expired",
                "level": "danger",
                "icon": "fa-ban",
                "title": "ĐÃ HẾT HẠN / QUÁ HẠN",
                "count": len(expired_svcs),
                "unit": "Hợp đồng",
                "breakdown": breakdown_services(expired_svcs),
                "action": "lug_phan_he.action_phan_he_service_expired",
            },
            {
                "id": "pay_due_soon",
                "level": "info",
                "icon": "fa-bell",
                "title": "SẮP ĐẾN HẠN THANH TOÁN",
                "count": len(due_soon_pays),
                "unit": "Hợp đồng",
                "breakdown": breakdown_payments(due_soon_pays),
                "action": "lug_phan_he.action_phan_he_payment_due_soon",
            },
        ]

        # --- Biểu đồ 12 tháng (theo năm filter) ---
        trend_months = []
        series = {m["id"]: [] for m in mien_meta}
        totals_row = []
        year_totals = {m["id"]: 0.0 for m in mien_meta}
        grand_total = 0.0

        for month in range(1, 13):
            m_start = fields.Date.to_date(f"{year}-{month:02d}-01")
            m_end = m_start + relativedelta(months=1, days=-1)
            cmap = cost_by_mien(services, m_start, m_end)
            month_total = sum(cmap.values())
            for m in mien_meta:
                val = cmap.get(m["id"], 0.0)
                series[m["id"]].append(val)
                year_totals[m["id"]] += val
            grand_total += month_total

            if month < today.month and year == today.year:
                status = "done"
                status_label = "Hoàn tất"
            elif month == today.month and year == today.year:
                status = "current"
                status_label = "Tháng hiện tại"
            elif year < today.year:
                status = "done"
                status_label = "Hoàn tất"
            elif year > today.year:
                status = "forecast"
                status_label = "Dự kiến"
            else:
                status = "forecast"
                status_label = "Dự kiến"

            totals_row.append({
                "month": month,
                "label": f"Tháng {month:02d}",
                "is_current": status == "current",
                "amounts": [
                    {"mien_id": m["id"], "amount": cmap.get(m["id"], 0.0), "color": m["color"]}
                    for m in mien_meta
                ],
                "total": month_total,
                "status": status,
                "status_label": status_label,
            })
            trend_months.append(f"T{month:02d}")

        max_trend = max(
            (v for vals in series.values() for v in vals),
            default=1,
        ) or 1

        trend_series = [
            {
                "id": m["id"],
                "name": m["name"],
                "short": m["short"],
                "color": m["color"],
                "values": series[m["id"]],
            }
            for m in mien_meta
        ]

        # Donut: cơ cấu tháng hiện tại
        cur_total = sum(cur_map.values()) or 1.0
        cost_structure = [
            {
                "id": m["id"],
                "name": m["name"],
                "short": m["short"],
                "amount": cur_map.get(m["id"], 0.0),
                "color": m["color"],
                "pct": round((cur_map.get(m["id"], 0.0) / cur_total) * 100, 1),
            }
            for m in mien_meta
            if cur_map.get(m["id"], 0.0) > 0
        ]
        if not cost_structure:
            cost_structure = [
                {
                    "id": m["id"],
                    "name": m["name"],
                    "short": m["short"],
                    "amount": 0.0,
                    "color": m["color"],
                    "pct": 0.0,
                }
                for m in mien_meta
            ]

        # Filter options
        miens_opts = [{"id": m["id"], "name": m["name"]} for m in mien_meta]
        area_domain = [("active", "=", True)]
        if mien_id:
            area_domain.append(("mien_id", "=", mien_id))
        areas = [
            {"id": a.id, "name": a.name, "mien_id": a.mien_id.id}
            for a in self.env["phan.he.area"].search(area_domain, order="name")
        ]
        employees = self.env["hr.employee"].search_read(
            [("id", "in", self.env["phan.he.store"].search([]).mapped("responsible_id").ids)],
            ["id", "name"],
            order="name",
        )

        return {
            "user_name": self.env.user.name,
            "role_label": self._role_label(),
            "updated_at": now.strftime("%H:%M %d/%m/%Y"),
            "app_title": filters.get("app_title") or "Quản lý dịch vụ",
            "service_type_code": service_type_code or "",
            "year": year,
            "current_month": today.month if year == today.year else 0,
            "date_from": fields.Date.to_string(date_from),
            "date_to": fields.Date.to_string(date_to),
            "mien_id": mien_id or "",
            "area_id": area_id or "",
            "employee_id": emp_id or "",
            "miens": miens_opts,
            "areas": areas,
            "employees": employees,
            "mien_meta": mien_meta,
            "region_kpis": region_kpis,
            "alert_cards": alert_cards,
            "alert_count": (
                len(expire_soon_svcs) + len(expired_svcs) + len(due_soon_pays)
            ),
            "expire_soon": len(expire_soon_svcs),
            "overdue_contract": len(expired_svcs),
            "trend_months": trend_months,
            "trend_series": trend_series,
            "max_trend": max_trend,
            "cost_structure": cost_structure,
            "monthly_table": totals_row,
            "year_totals": [
                {"mien_id": m["id"], "amount": year_totals[m["id"]], "color": m["color"]}
                for m in mien_meta
            ],
            "year_grand_total": grand_total,
            "currency_symbol": (
                self.env.ref("base.VND", raise_if_not_found=False)
                or self.env["res.currency"].search([("name", "=", "VND")], limit=1)
                or self.env.company.currency_id
            ).symbol or "đ",
        }

    @api.model
    def _delta_pct(self, current, previous):
        if not previous:
            return 100.0 if current else 0.0
        return round(((current - previous) / abs(previous)) * 100, 1)

    @api.model
    def export_monthly_cost_excel(self, filters=None):
        """Xuất bảng tổng hợp chi phí 12 tháng ra Excel."""
        import base64
        import io

        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError as exc:
            from odoo.exceptions import UserError
            raise UserError("Thiếu thư viện openpyxl trên server.") from exc

        data = self.get_dashboard_data(filters or {})
        year = data.get("year") or ""
        mien_meta = data.get("mien_meta") or []
        monthly = data.get("monthly_table") or []
        year_totals = data.get("year_totals") or []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Chi phi {year}"

        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill("solid", fgColor="E2E8F0")
        current_fill = PatternFill("solid", fgColor="FFF7ED")
        thin = Border(
            left=Side(style="thin", color="94A3B8"),
            right=Side(style="thin", color="94A3B8"),
            top=Side(style="thin", color="94A3B8"),
            bottom=Side(style="thin", color="94A3B8"),
        )
        right = Alignment(horizontal="right", vertical="center")
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")

        title = f"Bảng tổng hợp chi phí chi tiết từ tháng 1 đến tháng 12 (VNĐ · Năm {year})"
        headers = ["Tháng"] + [f"Chi phí {m.get('name') or ''}" for m in mien_meta] + [
            "Tổng chi phí",
            "Trạng thái / Lưu ý",
        ]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        cell = ws.cell(1, 1, title)
        cell.font = Font(bold=True, size=13, color="0F172A")
        cell.alignment = left

        for col, text in enumerate(headers, 1):
            c = ws.cell(3, col, text)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center if col == 1 or col == len(headers) else right
            c.border = thin

        row_idx = 4
        for row in monthly:
            values = [row.get("label") or ""]
            for amt in row.get("amounts") or []:
                values.append(float(amt.get("amount") or 0))
            values.append(float(row.get("total") or 0))
            values.append(row.get("status_label") or "")
            for col, val in enumerate(values, 1):
                c = ws.cell(row_idx, col, val)
                c.border = thin
                if col == 1:
                    c.alignment = left
                elif col == len(values):
                    c.alignment = center
                else:
                    c.alignment = right
                    c.number_format = "#,##0"
                if row.get("is_current"):
                    c.fill = current_fill
            row_idx += 1

        # Lũy kế
        foot = ["LŨY KẾ"]
        yt_map = {yt.get("mien_id"): float(yt.get("amount") or 0) for yt in year_totals}
        for m in mien_meta:
            foot.append(yt_map.get(m.get("id"), 0.0))
        foot.append(float(data.get("year_grand_total") or 0))
        foot.append("Tổng cả năm")
        for col, val in enumerate(foot, 1):
            c = ws.cell(row_idx, col, val)
            c.border = thin
            c.fill = total_fill
            c.font = Font(bold=True)
            if col == 1:
                c.alignment = left
            elif col == len(foot):
                c.alignment = center
            else:
                c.alignment = right
                c.number_format = "#,##0"

        ws.column_dimensions["A"].width = 14
        for i in range(2, len(headers)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
        ws.column_dimensions[openpyxl.utils.get_column_letter(len(headers))].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        return {
            "file_base64": base64.b64encode(buf.getvalue()).decode("ascii"),
            "filename": f"Bang_tong_hop_chi_phi_{year}.xlsx",
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    @api.model
    def get_payment_status_report(self, filters=None):
        """Báo cáo 4 danh sách thanh toán theo trạng thái."""
        filters = filters or {}
        today = fields.Date.context_today(self)

        year = today.year
        if filters.get("date_from"):
            year = fields.Date.to_date(filters["date_from"]).year
        date_from = fields.Date.to_date(filters.get("date_from") or f"{year}-01-01")
        date_to = fields.Date.to_date(filters.get("date_to") or f"{year}-12-31")

        mien_id = int(filters["mien_id"]) if filters.get("mien_id") else False
        area_id = int(filters["area_id"]) if filters.get("area_id") else False
        emp_id = int(filters["employee_id"]) if filters.get("employee_id") else False

        Service = self.env["phan.he.service"]
        Payment = self.env["phan.he.payment"]

        s_domain = [("active", "=", True)]
        service_type_code = (filters.get("service_type_code") or "").strip().lower()
        if service_type_code:
            stype = self.env["phan.he.service.type"].search(
                [("code", "=", service_type_code)], limit=1
            )
            if stype:
                s_domain.append(("service_type_id", "=", stype.id))
        if mien_id:
            s_domain.append(("mien_id", "=", mien_id))
        if area_id:
            s_domain.append(("area_id", "=", area_id))
        if emp_id:
            s_domain.append(("store_id.responsible_id", "=", emp_id))

        services = Service.search(s_domain)
        service_ids = services.ids or [0]

        def serialize(pays):
            rows = []
            for p in pays:
                rows.append({
                    "id": p.id,
                    "code": p.code or "",
                    "store": p.store_name or (p.store_id.name if p.store_id else "") or "",
                    "mien": p.mien_id.name if p.mien_id else "",
                    "provider": p.provider_id.name if p.provider_id else "",
                    "period": p.period or "",
                    "invoice_number": p.invoice_number or "",
                    "amount": p.amount or 0.0,
                    "date_due": fields.Date.to_string(p.date_due) if p.date_due else "",
                    "date_paid": fields.Date.to_string(p.date_paid) if p.date_paid else "",
                    "payment_state": p.payment_state,
                })
            return rows

        def search_state(states, order="date_due desc, id desc", limit=200):
            domain = [
                ("service_id", "in", service_ids),
                ("payment_state", "in", list(states)),
            ]
            # Lọc theo kỳ hạn / ngày TT trong khoảng năm đang xem
            domain += [
                "|",
                "&", ("date_due", ">=", date_from), ("date_due", "<=", date_to),
                "&", ("date_paid", ">=", date_from), ("date_paid", "<=", date_to),
            ]
            return Payment.search(domain, order=order, limit=limit)

        paid = search_state(("paid",), order="date_paid desc, id desc")
        due_soon = search_state(("due_soon",), order="date_due asc, id asc")
        pending = search_state(("pending", "not_due", "draft"), order="date_due asc, id asc")
        overdue = search_state(("overdue",), order="date_due asc, id asc")

        def bucket(key, title, tone, pays):
            total = sum(pays.mapped("amount"))
            return {
                "key": key,
                "title": title,
                "tone": tone,
                "count": len(pays),
                "total": total,
                "rows": serialize(pays),
            }

        tables = [
            bucket("paid", "Bảng 1: Danh sách đã thanh toán", "paid", paid),
            bucket("due_soon", "Bảng 2: Danh sách sắp thanh toán", "due_soon", due_soon),
            bucket("pending", "Bảng 3: Danh sách chờ thanh toán", "pending", pending),
            bucket("overdue", "Bảng 4: Danh sách quá hạn", "overdue", overdue),
        ]
        return {
            "date_from": fields.Date.to_string(date_from),
            "date_to": fields.Date.to_string(date_to),
            "tables": tables,
            "grand_count": sum(t["count"] for t in tables),
            "grand_total": sum(t["total"] for t in tables),
        }

    @api.model
    def _role_label(self):
        user = self.env.user
        if user.has_group("lug_phan_he.group_phan_he_admin"):
            return "Administrator"
        if user.has_group("lug_phan_he.group_phan_he_service_manager"):
            return "Service Manager"
        if user.has_group("lug_phan_he.group_phan_he_regional"):
            return "Regional Manager"
        if user.has_group("lug_phan_he.group_phan_he_area"):
            return "Area Manager"
        if user.has_group("lug_phan_he.group_phan_he_staff"):
            return "Staff"
        return "Viewer"
