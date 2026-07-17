# -*- coding: utf-8 -*-
"""Generate sample_data.xml from the Excel template."""
import re
import unicodedata
from datetime import datetime
from xml.sax.saxutils import escape

import openpyxl

PATH = r"c:\Users\nguye\Downloads\Bản sao của Bảng quản lý công việc 5.2.4.xlsx"
OUT = r"d:\Lap_odoo\odoo_time_off_custom\custom_addons\daily_work_task\data\sample_data.xml"

dept_map = {
    "Kinh Doanh": "kinh_doanh",
    "Marketing": "marketing",
    "Kỹ thuật": "ky_thuat",
    "IT": "it",
}
prio_map = {"Cao": "high", "Trung bình": "medium", "Thấp": "low"}
state_map = {
    "Chưa bắt đầu": "not_started",
    "Đang xử lý": "in_progress",
    "Đã hoàn thành": "done",
}


def slugify(name):
    n2 = unicodedata.normalize("NFD", name)
    n2 = "".join(c for c in n2 if unicodedata.category(c) != "Mn")
    n2 = n2.lower()
    n2 = re.sub(r"[^a-z0-9]+", "_", n2).strip("_")
    return n2 or "emp"


wb = openpyxl.load_workbook(PATH, data_only=True)
emp_sheet = wb["Danh sách nhân viên"]
emps = []
for r in range(2, emp_sheet.max_row + 1):
    name = emp_sheet.cell(r, 1).value
    email = emp_sheet.cell(r, 2).value
    if name:
        emps.append((name, email or "emailcuanhanvien@gmail.com"))

ws = wb["Danh sách công việc"]
tasks = []
assignees = {e[0] for e in emps}
for r in range(4, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if not name:
        continue
    deadline = ws.cell(r, 2).value
    dept = ws.cell(r, 3).value
    assignee = ws.cell(r, 4).value
    prio = ws.cell(r, 5).value
    state = ws.cell(r, 6).value
    note = ws.cell(r, 7).value or ""
    if assignee and assignee not in assignees:
        emps.append((assignee, "emailcuanhanvien@gmail.com"))
        assignees.add(assignee)
    if isinstance(deadline, datetime):
        deadline = deadline.strftime("%Y-%m-%d")
    tasks.append(
        {
            "name": name,
            "deadline": deadline,
            "dept": dept_map.get(dept, "it"),
            "assignee": assignee,
            "prio": prio_map.get(prio, "medium"),
            "state": state_map.get(state, "not_started"),
            "note": "" if note == "Nhập mô tả chi tiết" else note,
        }
    )

emp_ids = {}
lines = [
    '<?xml version="1.0" encoding="utf-8"?>',
    '<odoo noupdate="1">',
]
for name, email in emps:
    xid = "emp_" + slugify(name)
    emp_ids[name] = xid
    lines.append(f'    <record id="{xid}" model="daily.task.employee">')
    lines.append(f'        <field name="name">{escape(name)}</field>')
    lines.append(f'        <field name="email">{escape(email)}</field>')
    lines.append("    </record>")

for i, t in enumerate(tasks, 1):
    xid = f"task_sample_{i:03d}"
    lines.append(f'    <record id="{xid}" model="daily.task">')
    lines.append(f'        <field name="name">{escape(t["name"])}</field>')
    lines.append(f'        <field name="deadline">{t["deadline"]}</field>')
    lines.append(f'        <field name="department">{t["dept"]}</field>')
    lines.append(f'        <field name="assignee_id" ref="{emp_ids[t["assignee"]]}"/>')
    lines.append(f'        <field name="priority">{t["prio"]}</field>')
    lines.append(f'        <field name="state">{t["state"]}</field>')
    if t["note"]:
        lines.append(f'        <field name="note">{escape(t["note"])}</field>')
    lines.append("    </record>")

lines.append("</odoo>")
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")
print(f"Employees: {len(emps)}, Tasks: {len(tasks)}")
print(f"Written: {OUT}")
