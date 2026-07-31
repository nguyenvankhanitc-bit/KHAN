# -*- coding: utf-8 -*-

import base64
import io
import os
from calendar import monthrange

from dateutil.relativedelta import relativedelta

from odoo import api, fields, models, _
from odoo.exceptions import UserError


class PhanHePaymentFile(models.Model):
    _name = "phan.he.payment.file"
    _description = "File thanh toán Internet"
    _inherit = ["mail.thread", "mail.activity.mixin"]
    _order = "year desc, month desc, id desc"

    name = fields.Char(string="Tên file", required=True, tracking=True)
    code = fields.Char(
        string="Mã",
        copy=False,
        default=lambda self: self.env["ir.sequence"].next_by_code("phan.he.payment.file") or "New",
    )
    year = fields.Integer(
        string="Năm",
        required=True,
        default=lambda self: fields.Date.context_today(self).year,
        tracking=True,
    )
    month = fields.Integer(
        string="Tháng",
        required=True,
        default=lambda self: fields.Date.context_today(self).month,
        tracking=True,
    )
    period_label = fields.Char(string="Kỳ", compute="_compute_period_label", store=True)
    date_from = fields.Date(string="Từ ngày", compute="_compute_period_dates", store=True)
    date_to = fields.Date(string="Đến ngày", compute="_compute_period_dates", store=True)
    pay_before_days = fields.Integer(
        string="Thanh toán trước (ngày)",
        default=3,
        required=True,
        help="Ngày đề nghị thanh toán = Ngày đến hạn − số ngày này (mặc định 3 ngày).",
    )
    company_code = fields.Char(string="Mã công ty", default="ST")
    company_id = fields.Many2one(
        "res.company",
        string="Công ty",
        required=True,
        default=lambda self: self.env.company,
    )
    line_ids = fields.One2many("phan.he.payment.file.line", "file_id", string="Dòng thanh toán")
    line_count = fields.Integer(compute="_compute_totals")
    amount_total = fields.Monetary(
        string="Tổng thanh toán",
        compute="_compute_totals",
        currency_field="currency_id",
    )
    currency_id = fields.Many2one(
        "res.currency",
        default=lambda self: self.env.ref("base.VND", raise_if_not_found=False)
        or self.env.company.currency_id,
    )
    state = fields.Selection(
        selection=[
            ("draft", "Nháp"),
            ("ready", "Đã lấy dữ liệu"),
            ("done", "Hoàn tất"),
            ("cancel", "Hủy"),
        ],
        default="draft",
        tracking=True,
    )
    excel_file = fields.Binary(string="File Excel", attachment=True)
    excel_filename = fields.Char(string="Tên file Excel")
    note = fields.Text(string="Ghi chú")

    _period_uniq = models.Constraint(
        "unique(year, month, company_id)",
        "Đã có File thanh toán cho tháng/năm này.",
    )

    @api.depends("year", "month")
    def _compute_period_label(self):
        for rec in self:
            if rec.year and rec.month:
                rec.period_label = f"T{rec.month}/{rec.year}"
            else:
                rec.period_label = False

    @api.depends("year", "month")
    def _compute_period_dates(self):
        for rec in self:
            if rec.year and rec.month and 1 <= rec.month <= 12:
                last = monthrange(rec.year, rec.month)[1]
                rec.date_from = fields.Date.to_date(f"{rec.year}-{rec.month:02d}-01")
                rec.date_to = fields.Date.to_date(f"{rec.year}-{rec.month:02d}-{last:02d}")
            else:
                rec.date_from = False
                rec.date_to = False

    @api.depends("line_ids", "line_ids.amount")
    def _compute_totals(self):
        for rec in self:
            rec.line_count = len(rec.line_ids)
            rec.amount_total = sum(rec.line_ids.mapped("amount"))

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        today = fields.Date.context_today(self)
        year = res.get("year") or today.year
        month = res.get("month") or today.month
        if "name" in fields_list and not res.get("name"):
            res["name"] = f"Thanh toán internet T{month} {year}"
        return res

    @api.onchange("year", "month")
    def _onchange_period(self):
        if self.year and self.month:
            self.name = f"Thanh toán internet T{self.month} {self.year}"

    def action_generate_lines(self):
        """Lấy dữ liệu từ theo dõi Internet; ngày TT = hạn − pay_before_days."""
        self.ensure_one()
        if self.state == "done":
            raise UserError(_("File đã hoàn tất — mở lại Nháp nếu cần lấy lại dữ liệu."))
        if not self.date_from or not self.date_to:
            raise UserError(_("Tháng/năm không hợp lệ."))

        Service = self.env["phan.he.service"]
        internet = self.env.ref("lug_phan_he.service_type_internet", raise_if_not_found=False)
        domain = [
            ("active", "=", True),
            ("ops_status", "=", "active"),
            ("state", "not in", ("cancel", "liquidated")),
            "|",
            ("date_start", "=", False),
            ("date_start", "<=", self.date_to),
            "|",
            ("date_end", "=", False),
            ("date_end", ">=", self.date_from),
        ]
        if internet:
            domain.append(("service_type_id", "=", internet.id))

        services = Service.search(domain, order="mien_id, stt, store_id, id")
        before = self.pay_before_days or 3
        lines = []
        stt = 0
        for svc in services:
            # Ngày đến hạn: ưu tiên ngày TT tiếp theo trong/gần tháng file; không thì cuối tháng
            date_due = svc.next_payment_date
            if not date_due or date_due.year != self.year or date_due.month != self.month:
                date_due = self.date_to
            date_pay = date_due - relativedelta(days=before)

            stt += 1
            store_name = svc.store_id.name or ""
            # Cột NỘI DUNG để trống — kế toán/người đề nghị tự điền
            content = ""

            amount = svc.next_payment_amount or svc.contract_amount or 0.0
            account_info = svc.payment_info_text or ""
            if not account_info and svc.provider_id:
                bank = svc.provider_id.bank_account_ids.filtered("is_default")[:1] \
                    or svc.provider_id.bank_account_ids[:1]
                if bank:
                    account_info = (
                        f"- TÊN TK: {bank.account_name or ''}\n"
                        f"- STK: {bank.account_number or ''}\n"
                        f"- NGÂN HÀNG: {bank.bank_name or ''}"
                    )
                    if bank.transfer_content_template:
                        account_info += f"\n- Nội dung: {bank.transfer_content_template}"

            lines.append((0, 0, {
                "stt": stt,
                "company_code": self.company_code or "ST",
                "service_id": svc.id,
                "store_name": store_name,
                "customer_code": svc.customer_code or "",
                "content": content,
                "amount": amount,
                "account_info": account_info,
                "date_due": date_due,
                "date_pay": date_pay,
                "mien_id": svc.mien_id.id,
            }))

        self.line_ids.unlink()
        self.write({
            "line_ids": lines,
            "state": "ready" if lines else "draft",
            "name": self.name or f"Thanh toán internet T{self.month} {self.year}",
        })
        return True

    def action_confirm_done(self):
        for rec in self:
            if not rec.line_ids:
                raise UserError(_("Chưa có dòng thanh toán — bấm Lấy dữ liệu trước."))
            rec.state = "done"
        return True

    def action_reset_draft(self):
        self.write({"state": "draft"})
        return True

    def action_export_excel(self):
        """Xuất Excel đúng mẫu: tiêu đề, header cam nhạt, số tiền #,##0, tổng, chữ ký."""
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        except ImportError as exc:
            raise UserError(_("Thiếu thư viện openpyxl trên server.")) from exc

        wb = openpyxl.Workbook()
        ws = wb.active
        sheet_title = (self.period_label or f"T{self.month} {self.year}").replace("/", "-")
        sheet_title = "".join(ch for ch in sheet_title if ch not in r"\/*?:[]")[:31] or "ThanhToan"
        ws.title = sheet_title

        font_title = Font(name="Times New Roman", size=18, bold=True)
        font_header = Font(name="Times New Roman", size=10, bold=True)
        font_cell = Font(name="Times New Roman", size=10)
        font_cell_bold = Font(name="Times New Roman", size=10, bold=True)
        fill_peach = PatternFill("solid", fgColor="FBE4D5")
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Khớp mẫu: #,##0 + khoảng trống + ₫
        num_fmt = '_-* #,##0\\ _₫_-;\\-* #,##0\\ _₫_-;_-* "-"??\\ _₫_-;_-@'

        # --- Hàng 1: tiêu đề ---
        ws.merge_cells("A1:G1")
        ws.row_dimensions[1].height = 36
        title = f"DANH SÁCH THANH TOÁN INTERNET THÁNG {self.period_label or ''} "
        cell = ws["A1"]
        cell.value = title
        cell.font = font_title
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Hàng 2: trống ---
        ws.row_dimensions[2].height = 18

        # --- Hàng 3: header ---
        ws.row_dimensions[3].height = 30.75
        headers = ["STT", "CÔNG TY", "CỬA HÀNG", "MÃ KH", "NỘI DUNG", "SỐ TIỀN", "TÊN TÀI KHOẢN"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(3, col, h)
            cell.font = font_header
            cell.fill = fill_peach
            cell.border = thin
            cell.alignment = align_center
            if col == 6:
                cell.number_format = num_fmt

        # --- Dữ liệu ---
        lines = self.line_ids.sorted("stt")
        data_start = 4
        row_i = data_start
        for line in lines:
            acct = (line.account_info or "").strip()
            if acct.count("\n") >= 1:
                ws.row_dimensions[row_i].height = max(45, 18 * (acct.count("\n") + 2))
            values = [
                line.stt or 0,
                line.company_code or "ST",
                line.store_name or "",
                line.customer_code or "",
                "",  # NỘI DUNG để trống theo yêu cầu
                float(line.amount or 0),
                line.account_info or "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row_i, col, val)
                cell.border = thin
                if col in (1, 2):
                    cell.font = font_cell_bold
                    cell.alignment = align_center
                elif col == 6:
                    cell.font = font_cell
                    cell.alignment = align_center
                    cell.number_format = num_fmt
                elif col == 7:
                    cell.font = font_cell
                    cell.alignment = align_left
                    cell.number_format = "@"
                else:
                    cell.font = font_cell
                    cell.alignment = align_left
            row_i += 1

        data_end = row_i - 1 if lines else data_start - 1

        # --- Tổng ---
        total_row = row_i
        ws.row_dimensions[total_row].height = 31.5
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
        for col in range(1, 8):
            cell = ws.cell(total_row, col)
            cell.border = thin
            if col <= 6:
                cell.fill = fill_peach
            cell.font = font_header if col in (1, 6) else font_cell
            cell.alignment = align_center
        ws.cell(total_row, 1).value = "TỔNG THANH TOÁN"
        if lines:
            ws.cell(total_row, 6).value = f"=SUM(F{data_start}:F{data_end})"
        else:
            ws.cell(total_row, 6).value = 0
        ws.cell(total_row, 6).number_format = num_fmt
        ws.cell(total_row, 6).font = font_header

        # --- Chữ ký (cách 1 hàng trống như mẫu) ---
        sig_row = total_row + 2
        ws.merge_cells(start_row=sig_row, start_column=2, end_row=sig_row, end_column=4)
        ws.merge_cells(start_row=sig_row, start_column=5, end_row=sig_row, end_column=6)
        for col, label in ((2, "NGƯỜI ĐỀ NGHỊ"), (5, "KẾ TOÁN"), (7, "NGƯỜI DUYỆT")):
            cell = ws.cell(sig_row, col, label)
            cell.font = font_header
            cell.alignment = align_left

        # Ô trống để ký / ghi tên (cách ~6 hàng)
        name_row = sig_row + 6
        ws.merge_cells(start_row=name_row, start_column=2, end_row=name_row, end_column=4)
        ws.merge_cells(start_row=name_row, start_column=6, end_row=name_row, end_column=7)
        for col in (2, 6, 7):
            cell = ws.cell(name_row, col)
            cell.font = font_cell_bold
            cell.alignment = align_center if col != 7 else align_left

        # --- Độ rộng cột theo mẫu ---
        widths = {"A": 6.29, "B": 10.86, "C": 17.71, "D": 21.71, "E": 33.43, "F": 13.43, "G": 68.0}
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width

        # --- Logo từ mẫu ---
        static_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "description")
        logo0 = os.path.join(static_dir, "payment_logo_0.png")
        logo1 = os.path.join(static_dir, "payment_logo_1.png")
        try:
            if os.path.isfile(logo0):
                img0 = XLImage(logo0)
                img0.width = 228
                img0.height = 77
                ws.add_image(img0, "A1")
            if os.path.isfile(logo1):
                img1 = XLImage(logo1)
                img1.width = 50
                img1.height = 30
                ws.add_image(img1, f"B{sig_row}")
        except Exception:
            pass

        buf = io.BytesIO()
        wb.save(buf)
        data = base64.b64encode(buf.getvalue())
        filename = f"Thanh_toan_internet_T{self.month}_{self.year}.xlsx"
        self.write({"excel_file": data, "excel_filename": filename})

        Attachment = self.env["ir.attachment"].sudo()
        Attachment.search([
            ("res_model", "=", self._name),
            ("res_id", "=", self.id),
            ("name", "=", filename),
        ]).unlink()
        att = Attachment.create({
            "name": filename,
            "type": "binary",
            "datas": data,
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{att.id}?download=true",
            "target": "new",
        }

    def action_create_payments(self):
        """Tạo phiếu thanh toán (phan.he.payment) từ các dòng."""
        self.ensure_one()
        Payment = self.env["phan.he.payment"]
        created = self.env["phan.he.payment"]
        for line in self.line_ids.filtered(lambda l: l.service_id and not l.payment_id):
            pay = Payment.create({
                "service_id": line.service_id.id,
                "amount": line.amount,
                "date_due": line.date_due,
                "period": self.period_label,
                "payment_content": line.content,
                "payment_state": "pending",
                "provider_id": line.service_id.provider_id.id or False,
            })
            line.payment_id = pay.id
            created |= pay
        return {
            "type": "ir.actions.act_window",
            "name": _("Thanh toán đã tạo"),
            "res_model": "phan.he.payment",
            "view_mode": "list,form",
            "domain": [("id", "in", created.ids)],
        }

    @api.model
    def _cron_auto_generate_monthly(self):
        """Cuối mỗi tháng / đầu tháng: tạo file tháng hiện tại nếu chưa có và lấy dữ liệu."""
        today = fields.Date.context_today(self)
        company = self.env.company
        existing = self.search([
            ("year", "=", today.year),
            ("month", "=", today.month),
            ("company_id", "=", company.id),
        ], limit=1)
        if not existing:
            existing = self.create({
                "name": f"Thanh toán internet T{today.month} {today.year}",
                "year": today.year,
                "month": today.month,
                "company_id": company.id,
                "pay_before_days": 3,
            })
        if existing.state in ("draft", "ready"):
            existing.action_generate_lines()
        return True


class PhanHePaymentFileLine(models.Model):
    _name = "phan.he.payment.file.line"
    _description = "Dòng file thanh toán Internet"
    _order = "stt, id"

    file_id = fields.Many2one(
        "phan.he.payment.file", string="File thanh toán",
        required=True, ondelete="cascade", index=True,
    )
    stt = fields.Integer(string="STT")
    company_code = fields.Char(string="Công ty", default="ST")
    service_id = fields.Many2one("phan.he.service", string="Hợp đồng", ondelete="set null")
    store_name = fields.Char(string="Cửa hàng")
    customer_code = fields.Char(string="Mã KH")
    content = fields.Text(string="Nội dung")
    amount = fields.Monetary(string="Số tiền", currency_field="currency_id")
    currency_id = fields.Many2one(related="file_id.currency_id", store=True, readonly=True)
    account_info = fields.Text(string="Tên tài khoản")
    date_due = fields.Date(string="Ngày đến hạn")
    date_pay = fields.Date(
        string="Ngày TT đề nghị",
        help="Thanh toán trước hạn 3 ngày (hoặc theo cấu hình file).",
    )
    mien_id = fields.Many2one("phan.he.mien", string="Miền")
    payment_id = fields.Many2one("phan.he.payment", string="Phiếu TT", ondelete="set null")
